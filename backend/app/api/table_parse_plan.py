from __future__ import annotations

import io
import logging
from json import JSONDecodeError

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openai import APITimeoutError
from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.settings import Settings, get_settings
from app.services.table_parse_plan import PlanBasedTableParser

router = APIRouter(prefix="/table-parse-plan", tags=["table-parse-plan"])
logger = logging.getLogger(__name__)


@router.post("/parse")
async def parse_table_with_plan(
    file: UploadFile = File(...),
    sheet_name: str | None = Query(None),
    settings: Settings = Depends(get_settings),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Plan-based table parsing currently supports .xlsx and .xlsm files.",
        )

    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise HTTPException(status_code=404, detail=f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        parser = PlanBasedTableParser(settings)
        result = parser.parse_sheet(sheet)
    except HTTPException:
        raise
    except APITimeoutError as exc:
        logger.exception("Plan-based table parsing timed out")
        raise HTTPException(
            status_code=504,
            detail=(
                "LLM request timed out while creating the table parse plan. "
                "Increase LLM_TIMEOUT_SECONDS or try a smaller sheet."
            ),
        ) from exc
    except (JSONDecodeError, ValidationError) as exc:
        logger.exception("Plan-based table parsing returned an invalid plan")
        raise HTTPException(
            status_code=422,
            detail=f"The LLM returned an invalid TableParsePlan: {exc}",
        ) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Plan-based table parsing failed")
        raise HTTPException(
            status_code=500,
            detail=f"{exc.__class__.__name__}: {exc}",
        ) from exc

    coverage = result.coverage.model_dump()
    coverage["is_complete"] = result.coverage.is_complete

    return {
        "filename": filename,
        "sheet_name": sheet.title,
        "mode": "plan",
        "raw_plan_output": result.raw_plan_output,
        "parse_plan": result.parse_plan.model_dump(),
        "validation_warnings": result.validation_warnings,
        "coverage": coverage,
        "tree_with_cell_refs": result.tree_with_cell_refs,
        "tree": result.tree,
    }
