from __future__ import annotations

import json
import re
from copy import deepcopy
from dataclasses import dataclass
from json import JSONDecodeError
from typing import Any

import openpyxl
from langchain_core.messages import BaseMessage
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.core.settings import Settings
from app.prompts.table2tree import (
    FINAL_JSON_TREE_CONSTRUCTION,
    HEADER_ANALYSIS_PROMPT,
    HIERARCHY_VALUE_IDENTIFICATION_PROMPT,
    TABLE_SUMMARY_PROMPT,
)


@dataclass
class EnhancedTableParseResult:
    table_title: str
    markdown_table: str
    normalized_headers: str
    hierarchy_definition: str
    summary_text: str
    final_json_tree: str
    tree_with_cell_refs: dict[str, Any]
    tree: dict[str, Any]


def excel_to_markdown_with_cell_ref(sheet: Worksheet) -> str:
    merged_cells: dict[str, Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        merged_ref = (
            f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:"
            f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
        )
        merged_cells[merged_ref] = sheet.cell(
            merged_range.min_row, merged_range.min_col
        ).value

    table_data: list[list[str]] = []
    for row in range(1, sheet.max_row + 1):
        row_data: list[str] = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell_ref = f"{get_column_letter(col)}{row}"
            cell_value = None
            is_in_merged = False

            for merged_ref, value in merged_cells.items():
                start_ref, end_ref = merged_ref.split(":")
                start_col = openpyxl.utils.column_index_from_string(
                    "".join(filter(str.isalpha, start_ref))
                )
                start_row = int("".join(filter(str.isdigit, start_ref)))
                end_col = openpyxl.utils.column_index_from_string(
                    "".join(filter(str.isalpha, end_ref))
                )
                end_row = int("".join(filter(str.isdigit, end_ref)))

                if start_row <= row <= end_row and start_col <= col <= end_col:
                    is_in_merged = True
                    if row == start_row and col == start_col:
                        cell_value = f"{merged_ref} {value if value is not None else ''}"
                    break

            if not is_in_merged:
                cell_value = f"{cell_ref} {cell.value if cell.value is not None else ''}"

            row_data.append(cell_value if cell_value is not None else f"{cell_ref} ")
        table_data.append(row_data)

    if not table_data:
        return ""

    markdown_lines = [
        "| " + " | ".join(str(cell) for cell in table_data[0]) + " |",
        "| " + " | ".join(["---"] * sheet.max_column) + " |",
    ]
    markdown_lines.extend(
        "| " + " | ".join(str(cell) for cell in row) + " |"
        for row in table_data[1:]
    )
    return "\n".join(markdown_lines)


def extract_table_title(sheet: Worksheet, max_scan_rows: int = 5) -> str:
    for row in range(1, min(sheet.max_row, max_scan_rows) + 1):
        row_values: list[tuple[int, str]] = []
        for col in range(1, sheet.max_column + 1):
            text = _clean_cell_text(sheet.cell(row, col).value)
            if text:
                row_values.append((col, text))

        if not row_values:
            continue

        merged_title = _merged_row_title(sheet, row, row_values)
        if merged_title:
            return merged_title

        if row <= 2 and len(row_values) == 1:
            return row_values[0][1]

    return sheet.title


def _merged_row_title(
    sheet: Worksheet,
    row: int,
    row_values: list[tuple[int, str]],
) -> str:
    min_title_span = max(2, (sheet.max_column * 3 + 4) // 5)
    for col, text in row_values:
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
                and merged_range.max_col - merged_range.min_col + 1 >= min_title_span
            ):
                return text
    return ""


def _clean_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_json_with_merge(llm_output: str) -> dict[str, Any]:
    json_content = _extract_json_object(llm_output)

    def deep_merge(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
        merged = left.copy()
        for key, right_value in right.items():
            if key not in merged:
                merged[key] = right_value
                continue

            left_value = merged[key]
            if isinstance(left_value, dict) and isinstance(right_value, dict):
                merged[key] = deep_merge(left_value, right_value)
            else:
                if not isinstance(left_value, list):
                    merged[key] = [left_value]
                if isinstance(right_value, list):
                    merged[key].extend(right_value)
                else:
                    merged[key].append(right_value)
        return merged

    def recursive_merge_hook(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                if isinstance(result[key], dict) and isinstance(value, dict):
                    result[key] = deep_merge(result[key], value)
                else:
                    if not isinstance(result[key], list):
                        result[key] = [result[key]]
                    if isinstance(value, list):
                        result[key].extend(value)
                    else:
                        result[key].append(value)
            else:
                result[key] = value
        return result

    try:
        parsed = json.loads(json_content, object_pairs_hook=recursive_merge_hook)
    except JSONDecodeError:
        repaired_json = _repair_common_json_issues(json_content)
        parsed = json.loads(repaired_json, object_pairs_hook=recursive_merge_hook)
    if not isinstance(parsed, dict):
        raise ValueError("LLM final tree output must be a JSON object")
    return parsed


def convert_cell_positions_to_values(
    data_dict: dict[str, Any],
    sheet: Worksheet,
    range_policy: str = "merged_top_left",
) -> dict[str, Any]:
    result = deepcopy(data_dict)

    def recursive_convert(obj: Any) -> Any:
        if isinstance(obj, dict):
            return {key: recursive_convert(value) for key, value in obj.items()}
        if isinstance(obj, list):
            return [recursive_convert(item) for item in obj]
        if isinstance(obj, str):
            return get_value_from_ref(obj, sheet, range_policy=range_policy)
        return obj

    return recursive_convert(result)


def get_value_from_ref(ref: str, sheet: Worksheet, range_policy: str) -> Any:
    ref_u = ref.strip().upper()
    if _is_cell_position(ref_u):
        try:
            return sheet[ref_u].value
        except Exception:
            return ref

    if _is_cell_range(ref_u):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref_u)
            for merged_range in sheet.merged_cells.ranges:
                if range_boundaries(str(merged_range).upper()) == (
                    min_col,
                    min_row,
                    max_col,
                    max_row,
                ):
                    return sheet.cell(row=min_row, column=min_col).value

            cells = sheet[ref_u]
            if range_policy == "matrix":
                if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
                    return [[cell.value for cell in row] for row in cells]
                return [cell.value for cell in cells]
            if range_policy == "flatten":
                if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
                    return [cell.value for row in cells for cell in row]
                return [cell.value for cell in cells]
            return sheet.cell(row=min_row, column=min_col).value
        except Exception:
            return ref

    return ref


class LangChainEnhancedTableParser:
    def __init__(self, settings: Settings):
        if not settings.llm_model:
            raise ValueError("LLM_MODEL is required for enhanced table parsing")
        if not settings.llm_api_key:
            raise ValueError("LLM_API_KEY is required for enhanced table parsing")

        kwargs: dict[str, Any] = {
            "model": settings.llm_model,
            "api_key": settings.llm_api_key,
            "temperature": settings.llm_temperature,
            "timeout": settings.llm_timeout_seconds,
        }
        if settings.llm_base_url:
            kwargs["base_url"] = settings.llm_base_url

        self.llm = ChatOpenAI(**kwargs)

    def parse_sheet(self, sheet: Worksheet) -> EnhancedTableParseResult:
        table_title = extract_table_title(sheet)
        markdown_table = excel_to_markdown_with_cell_ref(sheet)
        normalized_headers = self._invoke_prompt(
            HEADER_ANALYSIS_PROMPT,
            TABLE_AS_JSON_STRING=markdown_table,
        )
        hierarchy_definition = self._invoke_prompt(
            HIERARCHY_VALUE_IDENTIFICATION_PROMPT,
            TABLE_AS_JSON_STRING=markdown_table,
            NORMALIZED_HEADERS_FROM_STEP_1=normalized_headers,
        )
        summary_text = self._invoke_prompt(
            TABLE_SUMMARY_PROMPT,
            TABLE_AS_JSON_STRING=markdown_table,
            NORMALIZED_HEADERS_FROM_STEP_1=normalized_headers,
            HIERARCHY_DEFINITION_FROM_STEP_2=hierarchy_definition,
        )
        final_json_tree = self._invoke_prompt(
            FINAL_JSON_TREE_CONSTRUCTION,
            TABLE_AS_JSON_STRING=markdown_table,
            NORMALIZED_HEADERS_FROM_STEP_1=normalized_headers,
            HIERARCHY_DEFINITION_FROM_STEP_2=hierarchy_definition,
        )
        tree_with_cell_refs = parse_json_with_merge(final_json_tree)
        tree = convert_cell_positions_to_values(tree_with_cell_refs, sheet)
        return EnhancedTableParseResult(
            table_title=table_title,
            markdown_table=markdown_table,
            normalized_headers=normalized_headers,
            hierarchy_definition=hierarchy_definition,
            summary_text=summary_text,
            final_json_tree=final_json_tree,
            tree_with_cell_refs=tree_with_cell_refs,
            tree=tree,
        )

    def _invoke_prompt(self, prompt_template: str, **kwargs: str) -> str:
        prompt = ChatPromptTemplate.from_template(prompt_template)
        response = (prompt | self.llm).invoke(kwargs)
        return _message_to_text(response)


def _message_to_text(message: BaseMessage | Any) -> str:
    content = getattr(message, "content", message)
    if isinstance(content, str):
        return _extract_after_think(content)
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text", "")))
            else:
                parts.append(str(item))
        return _extract_after_think("".join(parts))
    return _extract_after_think(str(content))


def _extract_after_think(response: str) -> str:
    if "</think>" in response:
        return response.split("</think>", 1)[1].strip()
    return response.strip()


def _extract_json_object(llm_output: str) -> str:
    if "```json" in llm_output:
        return llm_output.split("```json", 1)[1].split("```", 1)[0].strip()
    if "```" in llm_output:
        return llm_output.split("```", 1)[1].split("```", 1)[0].strip()

    start = llm_output.find("{")
    end = llm_output.rfind("}") + 1
    if start < 0 or end <= start:
        raise ValueError("No JSON object found in LLM output")
    return llm_output[start:end].strip()


def _repair_common_json_issues(json_content: str) -> str:
    # LLMs often leave trailing commas before a closing object/array.
    repaired = re.sub(r",(\s*[}\]])", r"\1", json_content)

    # If prose leaked around the object inside a fenced block, keep the outer object.
    start = repaired.find("{")
    end = repaired.rfind("}") + 1
    if start >= 0 and end > start:
        repaired = repaired[start:end]
    return repaired.strip()


def _is_cell_position(text: str) -> bool:
    return bool(re.match(r"^\$?[A-Z]+\$?[0-9]+$", text.strip().upper()))


def _is_cell_range(text: str) -> bool:
    return bool(
        re.match(r"^\$?[A-Z]+\$?[0-9]+:\$?[A-Z]+\$?[0-9]+$", text.strip().upper())
    )
