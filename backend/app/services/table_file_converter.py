from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


@dataclass
class ConvertedTableFile:
    original_filename: str
    normalized_filename: str
    source_content: bytes
    xlsx_content: bytes
    source_extension: str


SUPPORTED_TABLE_EXTENSIONS = {".csv", ".xls", ".xlsx", ".xlsm"}


def convert_table_file_to_xlsx(filename: str, content: bytes) -> ConvertedTableFile:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_TABLE_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_TABLE_EXTENSIONS))
        raise ValueError(f"Unsupported table file type: {extension}. Supported: {supported}")

    normalized_filename = f"{Path(filename).stem or 'uploaded-table'}.xlsx"
    if extension == ".xlsx":
        xlsx_content = _validate_xlsx(content)
    elif extension == ".xlsm":
        xlsx_content = _convert_xlsm_to_xlsx(content)
    elif extension == ".csv":
        xlsx_content = _convert_csv_to_xlsx(filename, content)
    else:
        xlsx_content = _convert_xls_to_xlsx(content)

    return ConvertedTableFile(
        original_filename=filename,
        normalized_filename=normalized_filename,
        source_content=content,
        xlsx_content=xlsx_content,
        source_extension=extension,
    )


def _validate_xlsx(content: bytes) -> bytes:
    try:
        load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid .xlsx file: {exc}") from exc
    return content


def _convert_xlsm_to_xlsx(content: bytes) -> bytes:
    try:
        workbook = load_workbook(io.BytesIO(content), keep_vba=False)
    except Exception as exc:
        raise ValueError(f"Invalid .xlsm file: {exc}") from exc

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _convert_csv_to_xlsx(filename: str, content: bytes) -> bytes:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin1"):
        try:
            frame = pd.read_csv(io.BytesIO(content), encoding=encoding)
            sheet_name = _safe_sheet_name(Path(filename).stem or "Sheet1")
            return _write_frames_to_xlsx({sheet_name: frame})
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except Exception as exc:
            raise ValueError(f"Failed to read CSV file: {exc}") from exc
    raise ValueError(f"Failed to decode CSV file: {last_error}")


def _convert_xls_to_xlsx(content: bytes) -> bytes:
    try:
        frames = pd.read_excel(io.BytesIO(content), sheet_name=None, engine="xlrd")
    except ImportError as exc:
        raise ValueError("Reading .xls files requires xlrd. Run: pip install xlrd==2.0.1") from exc
    except Exception as exc:
        raise ValueError(f"Failed to read .xls file: {exc}") from exc
    return _write_frames_to_xlsx(frames)


def _write_frames_to_xlsx(frames: dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for raw_sheet_name, frame in frames.items():
            frame.to_excel(writer, index=False, sheet_name=_safe_sheet_name(raw_sheet_name))
    return output.getvalue()


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", str(name).strip()) or "Sheet1"
    return cleaned[:31]
