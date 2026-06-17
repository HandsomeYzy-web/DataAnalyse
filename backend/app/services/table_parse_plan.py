# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from json import JSONDecodeError
from typing import Any

from langchain_core.messages import BaseMessage, HumanMessage
from langchain_openai import ChatOpenAI
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field, ValidationError, field_validator, model_validator

from app.core.settings import Settings


TABLE_PARSE_PLAN_PROMPT = """
你是一名电子表格结构解析专家。

你会收到一个 Excel 工作表的 JSON 表示。每个单元格都包含坐标、原始值、合并单元格后的有效值、合并区域、是否为空以及基础样式线索。

你的任务不是生成最终语义树，也不是逐格填写数据。你的任务是输出一个 TableParsePlan，让后续程序可以根据坐标确定性地构建语义树，并保证不遗漏任何值单元格。

请严格遵守以下规则：

1. 只输出一个合法 JSON 对象，不要输出 Markdown 代码块、解释文字或注释。
2. 表头、行头、分组名称必须保留原始文字，不要翻译、概括、改写或创造新名称。
3. 识别数据区时要依据表格结构、合并单元格、表头形态、数据分布和行列关系，不要依赖固定标记。例如“甲/乙”“1/2”“A/B”等只能作为辅助线索，不能作为必要条件。
4. value_columns 必须包含所有值列，即使该列的数据单元格全为空也必须包含。空列不能省略。
5. hierarchy_columns 表示用于构建行路径的列，例如名称、代码、类别、日期、地区、ID、重复行标签等。
6. 如果行头跨多个相邻列，必须按从左到右的顺序包含所有行头列。例如 A 列是父分类、B 列是具体条目，则 A、B 都应进入 hierarchy_columns。
7. value_columns 表示数据值列。如果多个值列共享同一个上级表头，请为它们设置相同的 group。
8. data_row_range 只能包含真实数据行，不能包含标题行、说明行、表头行、编号行、单位行、注释行或页脚。
9. 如果层级列中的空白单元格表示“沿用上一行/上一合并区域的值”，请将 hierarchy_fill_down 设置为 true。
10. 对复杂表格，必须为每一条数据行输出 row_paths。row_paths 是每一行完整的语义路径，用来表达父分类、合并行头、具体条目、明细行、限定行等结构。
11. 对“附属行/限定行/明细行”要挂到其所属的上一条具体条目下面，不能把重复的限定标签合并成全局节点。例如某行表示“博士生导师下面的 #女”，路径必须先包含“博士生导师”，再包含“#女”。
12. 如果当前行的某个 ref 指向合并单元格的子格，请使用该合并区域左上角的 ref 表达父级语义，避免把空白子格当作独立语义。
13. role 字段建议使用：
    - group：父级分组或大类
    - item：具体条目
    - qualifier：限定/附属行，例如“#女”“其中”“小计下的明细”等
    - code：代码列
    - note：必要的说明节点

输出 JSON 结构必须符合以下格式：
{
  "table_range": "A1:T21",
  "title_ranges": ["A1:T2"],
  "header_ranges": ["A3:T4"],
  "data_row_range": "A6:T21",
  "hierarchy_columns": [
    {"header": "指标名称", "col": "A"},
    {"header": "代码", "col": "B"}
  ],
  "value_columns": [
    {"header": "高职专科", "col": "C", "group": "高职专科"},
    {"header": "高职专科 - #女", "col": "D", "group": "高职专科"}
  ],
  "row_paths": [
    {
      "row": 10,
      "path": [
        {"header": "指标名称", "ref": "A10", "role": "group"},
        {"header": "指标名称", "ref": "B10", "role": "item"},
        {"header": "代码", "ref": "C10", "role": "code"}
      ]
    },
    {
      "row": 11,
      "path": [
        {"header": "指标名称", "ref": "A10", "role": "group"},
        {"header": "指标名称", "ref": "B10", "role": "item"},
        {"header": "指标名称", "ref": "B11", "role": "qualifier"},
        {"header": "代码", "ref": "C11", "role": "code"}
      ]
    }
  ],
  "hierarchy_fill_down": true,
  "notes": []
}

注意：
- row_paths 的数量应当覆盖 data_row_range 内的每一条数据行。
- row_paths 中的 ref 必须是单个 Excel 单元格坐标，例如 "A10"。
- value_columns 中的 col 必须是 Excel 列字母，例如 "C"。
- 不要在 JSON 中使用尾逗号。
- 不要输出 JSON 之外的任何内容。

[工作表 JSON]
{worksheet_json}
"""

TABLE_PARSE_PLAN_PROMPT = """
你是一名电子表格结构解析专家。
你会收到一个 Excel 工作表的 compact profile。profile 可能只包含采样行，但后续程序会在完整工作表上执行解析计划。
你的任务不是生成最终语义树，也不是逐格填写数据。你的任务是输出一个 TableParsePlan，让 Python 根据坐标规则扫描完整 sheet 并确定性构建树。

请严格遵守：
1. 只输出一个合法 JSON 对象，不要输出 Markdown 代码块、解释文字或注释。
2. 表头、行头、分组名称必须保留原始文字，不要翻译、概括、改写或创造新名称。
3. data_row_range 必须覆盖完整工作表中的真实数据行范围，不只覆盖采样行。
4. value_columns 必须包含所有值列，即使该列在采样行中全为空也要包含。
5. hierarchy_columns 表示用于构建完整行路径的列，例如名称、代码、类别、地区、年龄、ID 等。多列行头必须按从左到右包含。
6. 大表优先不要输出逐行 row_paths。只要可以用 hierarchy_columns + value_columns 规则覆盖完整数据区，就让 row_paths 为空。
7. 只有当某些采样行无法用列规则表达父子关系时，才为这些特殊行输出 row_paths；不要为每一行生成 row_paths。
8. ignored_rows 用于列出 data_row_range 内需要跳过的标题行、单位行、注释行、空白分隔行、小节标题行。不要把这些行当数据写树。
9. 如果层级列中的空白单元格表示沿用上一行或上一合并区域的值，请将 hierarchy_fill_down 设置为 true。
10. 如果输入 profile 显示一个 sheet 内有多个形状不同的区域，请选择主数据区域作为 table_range/data_row_range，并在 notes 说明其它区域需要另行切分；不要强行用一个 plan 覆盖多个异构区域。

输出 JSON 结构必须符合以下格式：
{
  "table_range": "A1:T21",
  "title_ranges": ["A1:T2"],
  "header_ranges": ["A3:T4"],
  "data_row_range": "A6:T21",
  "hierarchy_columns": [
    {"header": "指标名称", "col": "A"},
    {"header": "代码", "col": "B"}
  ],
  "value_columns": [
    {"header": "高职专科", "col": "C", "group": "招生数"},
    {"header": "高职专科 - #女", "col": "D", "group": "招生数"}
  ],
  "row_paths": [],
  "ignored_rows": [],
  "hierarchy_fill_down": true,
  "notes": []
}

注意：
- row_paths 可以为空。大表默认应为空，除非有少量特殊行需要手工路径。
- row_paths 中的 ref 必须是单个 Excel 单元格坐标，例如 "A10"。
- value_columns 中的 col 必须是 Excel 列字母，例如 "C"。
- ignored_rows 必须是行号数组，例如 [5, 9]。
- 不要在 JSON 中使用尾逗号。
- 不要输出 JSON 之外的任何内容。

[工作表 compact profile]
{worksheet_json}
"""


class SheetCell(BaseModel):
    coord: str
    row: int
    col: int
    col_letter: str
    value: Any = None
    effective_value: Any = None
    value_type: str
    merged_range: str | None = None
    is_merged_child: bool = False
    is_blank: bool = False
    number_format: str | None = None
    font_bold: bool = False
    horizontal_alignment: str | None = None


class SheetGrid(BaseModel):
    sheet_name: str
    max_row: int
    max_column: int
    table_range: str
    merged_ranges: list[str]
    rows: list[list[SheetCell]]


class HierarchyColumn(BaseModel):
    header: str
    col: str

    @field_validator("col")
    @classmethod
    def normalize_col(cls, value: str) -> str:
        value = value.strip().upper()
        if not re.fullmatch(r"[A-Z]+", value):
            raise ValueError("Column must be an Excel column letter")
        return value


class ValueColumn(BaseModel):
    header: str
    col: str
    group: str | None = None

    @field_validator("col")
    @classmethod
    def normalize_col(cls, value: str) -> str:
        value = value.strip().upper()
        if not re.fullmatch(r"[A-Z]+", value):
            raise ValueError("Column must be an Excel column letter")
        return value


class RowPathSegment(BaseModel):
    header: str
    ref: str | None = None
    value: Any = None
    role: str | None = None

    @field_validator("ref")
    @classmethod
    def normalize_ref(cls, value: str | None) -> str | None:
        if value is None:
            return value
        value = value.strip().upper()
        if not re.fullmatch(r"\$?[A-Z]+\$?[0-9]+", value):
            raise ValueError("ref must be a single Excel cell reference")
        return value

    @model_validator(mode="after")
    def require_ref_or_value(self) -> RowPathSegment:
        if self.ref is None and self.value is None:
            raise ValueError("Row path segment requires either ref or value")
        return self


class RowPath(BaseModel):
    row: int
    path: list[RowPathSegment]

    @model_validator(mode="after")
    def require_path(self) -> RowPath:
        if not self.path:
            raise ValueError("row path cannot be empty")
        return self


class TableParsePlan(BaseModel):
    table_range: str
    title_ranges: list[str] = Field(default_factory=list)
    header_ranges: list[str] = Field(default_factory=list)
    data_row_range: str
    hierarchy_columns: list[HierarchyColumn] = Field(default_factory=list)
    value_columns: list[ValueColumn]
    row_paths: list[RowPath] = Field(default_factory=list)
    ignored_rows: list[int] = Field(default_factory=list)
    hierarchy_fill_down: bool = True
    notes: list[str] = Field(default_factory=list)

    @model_validator(mode="after")
    def require_columns(self) -> TableParsePlan:
        if not self.hierarchy_columns and not self.row_paths:
            raise ValueError("Either hierarchy_columns or row_paths must be provided")
        if not self.value_columns:
            raise ValueError("value_columns cannot be empty")
        return self


class CoverageReport(BaseModel):
    data_rows: int
    value_columns: int
    expected_cells: int
    covered_cells: int
    missing_cells: list[str] = Field(default_factory=list)
    skipped_rows: list[int] = Field(default_factory=list)

    @property
    def is_complete(self) -> bool:
        return self.expected_cells == self.covered_cells and not self.missing_cells


class PlanBuildResult(BaseModel):
    raw_plan_output: str
    parse_plan: TableParsePlan
    validation_warnings: list[str]
    tree_with_cell_refs: dict[str, Any]
    tree: dict[str, Any]
    coverage: CoverageReport


def extract_sheet_grid(sheet: Worksheet) -> SheetGrid:
    merged_lookup = _build_merged_lookup(sheet)
    rows: list[list[SheetCell]] = []

    for row_idx in range(1, sheet.max_row + 1):
        row_cells: list[SheetCell] = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            coord = f"{get_column_letter(col_idx)}{row_idx}"
            merged_info = merged_lookup.get((row_idx, col_idx))
            merged_range = merged_info["range"] if merged_info else None
            is_merged_child = bool(merged_info and not merged_info["is_top_left"])
            effective_value = (
                merged_info["top_left_value"] if merged_info else cell.value
            )

            row_cells.append(
                SheetCell(
                    coord=coord,
                    row=row_idx,
                    col=col_idx,
                    col_letter=get_column_letter(col_idx),
                    value=_json_safe(cell.value),
                    effective_value=_json_safe(effective_value),
                    value_type=type(effective_value).__name__,
                    merged_range=merged_range,
                    is_merged_child=is_merged_child,
                    is_blank=_is_blank(effective_value),
                    number_format=cell.number_format,
                    font_bold=bool(cell.font and cell.font.bold),
                    horizontal_alignment=cell.alignment.horizontal,
                )
            )
        rows.append(row_cells)

    return SheetGrid(
        sheet_name=sheet.title,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        table_range=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
        merged_ranges=[str(merged_range) for merged_range in sheet.merged_cells.ranges],
        rows=rows,
    )


def grid_to_llm_json(grid: SheetGrid) -> str:
    sample_row_indexes = _sample_row_indexes(grid)
    payload: dict[str, Any] = {
        "sheet_name": grid.sheet_name,
        "max_row": grid.max_row,
        "max_column": grid.max_column,
        "table_range": grid.table_range,
        "merged_ranges": grid.merged_ranges,
        "sampled_rows": [
            _compact_row(grid.rows[row_idx - 1])
            for row_idx in sample_row_indexes
            if 1 <= row_idx <= len(grid.rows)
        ],
        "row_profiles": [_row_profile(row) for row in grid.rows],
        "column_profiles": _column_profiles(grid),
    }
    return json.dumps(payload, ensure_ascii=False)


def _sample_row_indexes(grid: SheetGrid) -> list[int]:
    row_count = grid.max_row
    selected: set[int] = set()
    selected.update(range(1, min(row_count, 20) + 1))
    selected.update(range(max(1, row_count - 8), row_count + 1))

    if row_count > 28:
        step = max(1, row_count // 12)
        for row_idx in range(21, row_count - 8, step):
            selected.add(row_idx)
            if row_idx + 1 <= row_count:
                selected.add(row_idx + 1)

    for row in grid.rows:
        if any(cell.merged_range and not cell.is_merged_child for cell in row):
            selected.add(row[0].row)
    return sorted(row_idx for row_idx in selected if 1 <= row_idx <= row_count)


def _compact_row(row: list[SheetCell]) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    for cell in row:
        if cell.is_merged_child:
            continue
        value = cell.effective_value
        if cell.is_blank and not cell.merged_range:
            continue
        item: dict[str, Any] = {
            "c": cell.coord,
            "v": _compact_value(value),
        }
        if cell.merged_range:
            item["m"] = cell.merged_range
            if cell.is_merged_child:
                item["mc"] = True
        if cell.font_bold:
            item["b"] = True
        if cell.horizontal_alignment:
            item["a"] = cell.horizontal_alignment
        cells.append(item)
    return {
        "row": row[0].row if row else None,
        "cells": cells,
    }


def _row_profile(row: list[SheetCell]) -> dict[str, Any]:
    non_blank = [cell for cell in row if not cell.is_blank]
    numeric_count = sum(1 for cell in non_blank if isinstance(cell.effective_value, (int, float)))
    text_count = len(non_blank) - numeric_count
    first_values = [
        {"c": cell.coord, "v": _compact_value(cell.effective_value)}
        for cell in non_blank[:4]
    ]
    return {
        "row": row[0].row if row else None,
        "non_blank": len(non_blank),
        "numeric": numeric_count,
        "text": text_count,
        "first_values": first_values,
    }


def _column_profiles(grid: SheetGrid) -> list[dict[str, Any]]:
    profiles: list[dict[str, Any]] = []
    for col_idx in range(1, grid.max_column + 1):
        cells = [row[col_idx - 1] for row in grid.rows if len(row) >= col_idx]
        non_blank = [cell for cell in cells if not cell.is_blank]
        numeric_count = sum(1 for cell in non_blank if isinstance(cell.effective_value, (int, float)))
        text_count = len(non_blank) - numeric_count
        first_values = [
            {"c": cell.coord, "v": _compact_value(cell.effective_value)}
            for cell in non_blank[:5]
        ]
        profiles.append(
            {
                "col": get_column_letter(col_idx),
                "non_blank": len(non_blank),
                "numeric": numeric_count,
                "text": text_count,
                "first_values": first_values,
            }
        )
    return profiles


def _compact_value(value: Any, max_length: int = 80) -> Any:
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value).strip()
        return text[:max_length] + "..." if len(text) > max_length else text
    return value


class PlanBasedTableParser:
    def __init__(self, settings: Settings):
        if not settings.llm_model:
            raise ValueError("LLM_MODEL is required for plan-based table parsing")
        if not settings.llm_api_key:
            raise ValueError("LLM_API_KEY is required for plan-based table parsing")

        kwargs: dict[str, Any] = {
            "model": settings.llm_model,
            "api_key": settings.llm_api_key,
            "temperature": settings.llm_temperature,
            "timeout": settings.llm_timeout_seconds,
        }
        if settings.llm_base_url:
            kwargs["base_url"] = settings.llm_base_url
        self.llm = ChatOpenAI(**kwargs)

    def parse_sheet(self, sheet: Worksheet) -> PlanBuildResult:
        grid = extract_sheet_grid(sheet)
        raw_plan_output = self.generate_parse_plan(grid)
        parse_plan = parse_table_parse_plan(raw_plan_output)
        parse_plan = normalize_parse_plan(parse_plan, sheet)
        validation_warnings = validate_parse_plan(parse_plan, sheet)
        tree_refs, tree, coverage = build_tree_from_plan(sheet, parse_plan)
        return PlanBuildResult(
            raw_plan_output=raw_plan_output,
            parse_plan=parse_plan,
            validation_warnings=validation_warnings,
            tree_with_cell_refs=tree_refs,
            tree=tree,
            coverage=coverage,
        )

    def generate_parse_plan(self, grid: SheetGrid) -> str:
        prompt_text = TABLE_PARSE_PLAN_PROMPT.replace(
            "{worksheet_json}",
            grid_to_llm_json(grid),
        )
        response = self.llm.invoke([HumanMessage(content=prompt_text)])
        return _message_to_text(response)


def parse_table_parse_plan(llm_output: str) -> TableParsePlan:
    json_content = _extract_json_object(llm_output)
    try:
        parsed = json.loads(json_content)
    except JSONDecodeError as first_exc:
        repaired_json = _repair_common_json_issues(json_content)
        try:
            parsed = json.loads(repaired_json)
        except JSONDecodeError:
            print(
                "\n[LLM_PARSE_PLAN_JSON_PARSE_FAILED]\n"
                f"first_error={first_exc}\n"
                f"raw_output:\n{llm_output}\n"
                f"extracted_json:\n{json_content}\n"
                f"repaired_json:\n{repaired_json}\n"
                "[/LLM_PARSE_PLAN_JSON_PARSE_FAILED]\n",
                flush=True,
            )
            raise
    try:
        return TableParsePlan.model_validate(parsed)
    except ValidationError:
        print(
            "\n[LLM_PARSE_PLAN_VALIDATION_FAILED]\n"
            f"raw_output:\n{llm_output}\n"
            f"parsed_json:\n{json.dumps(parsed, ensure_ascii=False, indent=2, default=str)}\n"
            "[/LLM_PARSE_PLAN_VALIDATION_FAILED]\n",
            flush=True,
        )
        raise


def normalize_parse_plan(plan: TableParsePlan, sheet: Worksheet) -> TableParsePlan:
    """Fill obvious structural gaps in the plan without using locale-specific tokens."""
    if plan.row_paths:
        return plan.model_copy(update={"row_paths": _repair_row_paths(plan.row_paths, sheet)})
    if not plan.value_columns:
        return plan

    _, data_min_row, _, data_max_row = range_boundaries(plan.data_row_range)
    table_min_col, _, _, _ = range_boundaries(plan.table_range)
    first_value_col = min(column_index_from_string(col.col) for col in plan.value_columns)
    hierarchy_col_indexes = {
        column_index_from_string(col.col) for col in plan.hierarchy_columns
    }
    value_col_indexes = {column_index_from_string(col.col) for col in plan.value_columns}

    hierarchy_columns = list(plan.hierarchy_columns)
    for col_idx in range(table_min_col, first_value_col):
        if col_idx in hierarchy_col_indexes or col_idx in value_col_indexes:
            continue
        if not _column_has_data_value(sheet, col_idx, data_min_row, data_max_row):
            continue

        hierarchy_columns.append(
            HierarchyColumn(
                header=_infer_header_for_column(sheet, col_idx, plan),
                col=get_column_letter(col_idx),
            )
        )

    hierarchy_columns.sort(key=lambda item: column_index_from_string(item.col))
    return plan.model_copy(update={"hierarchy_columns": hierarchy_columns})


def validate_parse_plan(plan: TableParsePlan, sheet: Worksheet) -> list[str]:
    warnings: list[str] = []
    _assert_range_inside_sheet(plan.table_range, sheet, "table_range")
    _assert_range_inside_sheet(plan.data_row_range, sheet, "data_row_range")

    for idx, range_ref in enumerate(plan.title_ranges):
        _assert_range_inside_sheet(range_ref, sheet, f"title_ranges[{idx}]")
    for idx, range_ref in enumerate(plan.header_ranges):
        _assert_range_inside_sheet(range_ref, sheet, f"header_ranges[{idx}]")

    table_min_col, _, table_max_col, _ = range_boundaries(plan.table_range)
    data_min_col, data_min_row, data_max_col, data_max_row = range_boundaries(
        plan.data_row_range
    )
    if data_min_row > data_max_row:
        raise ValueError("data_row_range has no rows")
    if data_min_col < table_min_col or data_max_col > table_max_col:
        warnings.append("data_row_range extends outside table_range columns")

    seen_value_cols: set[str] = set()
    data_rows = set(range(data_min_row, data_max_row + 1))
    seen_row_paths: set[int] = set()

    for row_idx in plan.ignored_rows:
        if row_idx not in data_rows:
            warnings.append(f"ignored_rows contains row outside data_row_range: {row_idx}")

    for row_path in plan.row_paths:
        if row_path.row not in data_rows:
            raise ValueError(f"row_paths contains row outside data_row_range: {row_path.row}")
        if row_path.row in seen_row_paths:
            warnings.append(f"Duplicate row_path for row: {row_path.row}")
        seen_row_paths.add(row_path.row)
        for segment in row_path.path:
            if segment.ref:
                _assert_cell_ref_inside_sheet(segment.ref, sheet, "row_paths.ref")

    if plan.row_paths and seen_row_paths != data_rows:
        missing_rows = sorted(data_rows - seen_row_paths)
        warnings.append(f"row_paths missing rows: {missing_rows}")

    for column in [*plan.hierarchy_columns, *plan.value_columns]:
        col_idx = column_index_from_string(column.col)
        if col_idx < table_min_col or col_idx > table_max_col:
            raise ValueError(f"Column {column.col} is outside table_range")
        if col_idx > sheet.max_column:
            raise ValueError(f"Column {column.col} is outside the worksheet")

    for value_column in plan.value_columns:
        if value_column.col in seen_value_cols:
            warnings.append(f"Duplicate value column: {value_column.col}")
        seen_value_cols.add(value_column.col)

    return warnings


def build_tree_from_plan(
    sheet: Worksheet,
    plan: TableParsePlan,
) -> tuple[dict[str, Any], dict[str, Any], CoverageReport]:
    _, data_min_row, _, data_max_row = range_boundaries(plan.data_row_range)

    tree_refs: dict[str, Any] = {}
    missing_cells: list[str] = []
    skipped_rows: list[int] = []
    covered_cells = 0
    last_hierarchy_values: dict[str, str] = {}
    row_path_map = {row_path.row: row_path.path for row_path in plan.row_paths}
    ignored_rows = set(plan.ignored_rows)

    for row_idx in range(data_min_row, data_max_row + 1):
        if row_idx in ignored_rows:
            skipped_rows.append(row_idx)
            continue

        if row_path_map:
            segments = row_path_map.get(row_idx)
            if not segments:
                skipped_rows.append(row_idx)
                for value_column in plan.value_columns:
                    missing_cells.append(f"{value_column.col}{row_idx}")
                continue
            current, row_path_parts = _build_current_from_row_path(
                tree_refs,
                sheet,
                segments,
            )
        else:
            current = tree_refs
            row_path_parts = 0
            previous_hierarchy_value = ""

            for hierarchy_column in plan.hierarchy_columns:
                col_idx = column_index_from_string(hierarchy_column.col)
                raw_value = get_effective_cell_value(sheet, row_idx, col_idx)
                display_value = _display_value(raw_value)

                if not display_value and plan.hierarchy_fill_down:
                    display_value = last_hierarchy_values.get(hierarchy_column.col, "")
                elif display_value:
                    last_hierarchy_values[hierarchy_column.col] = display_value

                if not display_value:
                    continue
                if (
                    previous_hierarchy_value
                    and display_value == previous_hierarchy_value
                    and _is_merged_child_cell(sheet, row_idx, col_idx)
                ):
                    continue

                semantic_key = f"{hierarchy_column.header} - {display_value}"
                current = current.setdefault(semantic_key, {})
                previous_hierarchy_value = display_value
                row_path_parts += 1

        if row_path_parts == 0:
            skipped_rows.append(row_idx)
            for value_column in plan.value_columns:
                missing_cells.append(f"{value_column.col}{row_idx}")
            continue

        for value_column in plan.value_columns:
            col_idx = column_index_from_string(value_column.col)
            ref = f"{get_column_letter(col_idx)}{row_idx}"
            if value_column.group:
                group_bucket = current.setdefault(value_column.group, {})
                group_bucket[value_column.header] = ref
            else:
                current[value_column.header] = ref
            covered_cells += 1

    data_rows = max(0, data_max_row - data_min_row + 1 - len(
        [row_idx for row_idx in ignored_rows if data_min_row <= row_idx <= data_max_row]
    ))
    expected_cells = data_rows * len(plan.value_columns)
    coverage = CoverageReport(
        data_rows=data_rows,
        value_columns=len(plan.value_columns),
        expected_cells=expected_cells,
        covered_cells=covered_cells,
        missing_cells=missing_cells,
        skipped_rows=skipped_rows,
    )
    return tree_refs, convert_refs_to_values(tree_refs, sheet), coverage


def convert_refs_to_values(data: dict[str, Any], sheet: Worksheet) -> dict[str, Any]:
    def convert(value: Any) -> Any:
        if isinstance(value, dict):
            return {key: convert(item) for key, item in value.items()}
        if isinstance(value, list):
            return [convert(item) for item in value]
        if isinstance(value, str) and _is_cell_ref(value):
            min_col, min_row, _, _ = range_boundaries(value)
            return _json_safe(get_effective_cell_value(sheet, min_row, min_col))
        return value

    return convert(data)


def get_effective_cell_value(sheet: Worksheet, row_idx: int, col_idx: int) -> Any:
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return sheet.cell(min_row, min_col).value
    return sheet.cell(row_idx, col_idx).value


def _build_current_from_row_path(
    tree_refs: dict[str, Any],
    sheet: Worksheet,
    segments: list[RowPathSegment],
) -> tuple[dict[str, Any], int]:
    current = tree_refs
    row_path_parts = 0
    previous_key = ""
    for segment in segments:
        display_value = _display_value(segment.value)
        if not display_value and segment.ref:
            col_idx, row_idx = _cell_ref_to_indexes(segment.ref)
            display_value = _display_value(get_effective_cell_value(sheet, row_idx, col_idx))
        if not display_value:
            continue

        semantic_key = f"{segment.header} - {display_value}"
        if semantic_key == previous_key:
            continue
        current = current.setdefault(semantic_key, {})
        previous_key = semantic_key
        row_path_parts += 1
    return current, row_path_parts


def _repair_row_paths(row_paths: list[RowPath], sheet: Worksheet) -> list[RowPath]:
    repaired_paths: list[RowPath] = []
    previous_parent_segments: list[RowPathSegment] = []

    for row_path in sorted(row_paths, key=lambda item: item.row):
        segments = _deduplicate_row_path_segments(row_path.path, sheet)
        qualifier_index = _first_qualifier_index(segments)

        if qualifier_index is not None and previous_parent_segments:
            prefix = [
                segment
                for segment in segments[:qualifier_index]
                if not _is_code_segment(segment) and not _is_qualifier_segment(segment)
            ]
            missing_parent_tail = _missing_parent_tail(
                prefix,
                previous_parent_segments,
                sheet,
            )
            if missing_parent_tail:
                segments = [
                    *segments[:qualifier_index],
                    *missing_parent_tail,
                    *segments[qualifier_index:],
                ]

        repaired_paths.append(row_path.model_copy(update={"path": segments}))

        if qualifier_index is None:
            parent_segments = [
                segment
                for segment in segments
                if not _is_code_segment(segment) and not _is_qualifier_segment(segment)
            ]
            if parent_segments:
                previous_parent_segments = parent_segments

    return repaired_paths


def _deduplicate_row_path_segments(
    segments: list[RowPathSegment],
    sheet: Worksheet,
) -> list[RowPathSegment]:
    deduped: list[RowPathSegment] = []
    for segment in segments:
        if deduped and _same_semantic_segment(deduped[-1], segment, sheet):
            if _is_qualifier_segment(segment) and not _is_qualifier_segment(deduped[-1]):
                deduped[-1] = deduped[-1].model_copy(update={"role": segment.role})
            continue
        deduped.append(segment)
    return deduped


def _first_qualifier_index(segments: list[RowPathSegment]) -> int | None:
    for index, segment in enumerate(segments):
        if _is_qualifier_segment(segment):
            return index
    return None


def _missing_parent_tail(
    prefix: list[RowPathSegment],
    previous_parent_segments: list[RowPathSegment],
    sheet: Worksheet,
) -> list[RowPathSegment]:
    if not prefix:
        return previous_parent_segments
    if len(prefix) > len(previous_parent_segments):
        return []
    for index, segment in enumerate(prefix):
        if not _same_semantic_segment(segment, previous_parent_segments[index], sheet):
            return []
    return previous_parent_segments[len(prefix):]


def _same_semantic_segment(
    left: RowPathSegment,
    right: RowPathSegment,
    sheet: Worksheet,
) -> bool:
    return (
        left.header == right.header
        and _segment_display_value(left, sheet) == _segment_display_value(right, sheet)
    )


def _segment_display_value(segment: RowPathSegment, sheet: Worksheet) -> str:
    display_value = _display_value(segment.value)
    if display_value or not segment.ref:
        return display_value
    col_idx, row_idx = _cell_ref_to_indexes(segment.ref)
    return _display_value(get_effective_cell_value(sheet, row_idx, col_idx))


def _is_code_segment(segment: RowPathSegment) -> bool:
    role = (segment.role or "").strip().lower()
    return role == "code"


def _is_qualifier_segment(segment: RowPathSegment) -> bool:
    role = (segment.role or "").strip().lower()
    return role in {"qualifier", "detail", "subitem", "sub_item", "modifier"}


def _column_has_data_value(
    sheet: Worksheet,
    col_idx: int,
    data_min_row: int,
    data_max_row: int,
) -> bool:
    for row_idx in range(data_min_row, data_max_row + 1):
        if _display_value(get_effective_cell_value(sheet, row_idx, col_idx)):
            return True
    return False


def _infer_header_for_column(sheet: Worksheet, col_idx: int, plan: TableParsePlan) -> str:
    for header_range in plan.header_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(header_range)
        if not min_col <= col_idx <= max_col:
            continue
        for row_idx in range(min_row, max_row + 1):
            value = _display_value(get_effective_cell_value(sheet, row_idx, col_idx))
            if value:
                return value
    return f"Column {get_column_letter(col_idx)}"


def _is_merged_child_cell(sheet: Worksheet, row_idx: int, col_idx: int) -> bool:
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return not (row_idx == min_row and col_idx == min_col)
    return False


def _build_merged_lookup(sheet: Worksheet) -> dict[tuple[int, int], dict[str, Any]]:
    lookup: dict[tuple[int, int], dict[str, Any]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_value = sheet.cell(min_row, min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                lookup[(row_idx, col_idx)] = {
                    "range": str(merged_range),
                    "top_left_value": top_left_value,
                    "is_top_left": row_idx == min_row and col_idx == min_col,
                }
    return lookup


def _assert_range_inside_sheet(range_ref: str, sheet: Worksheet, field_name: str) -> None:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except ValueError as exc:
        raise ValueError(f"{field_name} is not a valid Excel range: {range_ref}") from exc

    if min_row < 1 or min_col < 1:
        raise ValueError(f"{field_name} starts outside the worksheet")
    if max_row > sheet.max_row or max_col > sheet.max_column:
        raise ValueError(f"{field_name} exceeds worksheet bounds: {range_ref}")


def _assert_cell_ref_inside_sheet(ref: str, sheet: Worksheet, field_name: str) -> None:
    col_idx, row_idx = _cell_ref_to_indexes(ref)
    if row_idx < 1 or col_idx < 1 or row_idx > sheet.max_row or col_idx > sheet.max_column:
        raise ValueError(f"{field_name} exceeds worksheet bounds: {ref}")


def _cell_ref_to_indexes(ref: str) -> tuple[int, int]:
    min_col, min_row, _, _ = range_boundaries(ref)
    return min_col, min_row


def _message_to_text(message: BaseMessage | Any) -> str:
    content = getattr(message, "content", message)
    if isinstance(content, str):
        return _extract_after_think(content)
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text", "")))
            else:
                parts.append(str(item))
        return _extract_after_think("".join(parts))
    return _extract_after_think(str(content))


def _extract_after_think(response: str) -> str:
    if "</think>" in response:
        return response.split("</think>", 1)[1].strip()
    return response.strip()


def _extract_json_object(llm_output: str) -> str:
    if "```json" in llm_output:
        return llm_output.split("```json", 1)[1].split("```", 1)[0].strip()
    if "```" in llm_output:
        return llm_output.split("```", 1)[1].split("```", 1)[0].strip()

    start = llm_output.find("{")
    end = llm_output.rfind("}") + 1
    if start < 0 or end <= start:
        raise ValueError("No JSON object found in LLM output")
    return llm_output[start:end].strip()


def _repair_common_json_issues(json_content: str) -> str:
    repaired = re.sub(r",(\s*[}\]])", r"\1", json_content)
    start = repaired.find("{")
    end = repaired.rfind("}") + 1
    if start >= 0 and end > start:
        repaired = repaired[start:end]
    return repaired.strip()


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    return _display_value(value) == ""


def _is_cell_ref(value: str) -> bool:
    return bool(re.fullmatch(r"\$?[A-Z]+\$?[0-9]+", value.strip().upper()))
