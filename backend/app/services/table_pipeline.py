from __future__ import annotations

import io
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any
from uuid import uuid4

from langchain_core.messages import HumanMessage
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from openpyxl import Workbook, load_workbook

from app.core.settings import Settings
from app.prompts.table2tree import TABLE_SUMMARY_PROMPT
from app.services.table2tree_enhanced import LangChainEnhancedTableParser
from app.services.table_file_converter import convert_table_file_to_xlsx
from app.services.table_object_storage import TableObjectStorage
from app.services.table_parse_plan import PlanBasedTableParser
from app.services.table_qa import TableQAService
from app.services.table_search_index import TableSearchIndex, build_table_summary


MAX_PARALLEL_SUB_QUERIES = 4
MAX_REWRITTEN_QUERIES = 8


@dataclass
class PipelineParseResult:
    parse_mode: str
    large_table_reason: str | None
    table_title: str
    markdown_table: str
    normalized_headers: str
    hierarchy_definition: str
    summary_text: str
    final_json_tree: str | None
    tree_with_cell_refs: dict[str, Any]
    tree: dict[str, Any]
    parse_plan: dict[str, Any] | None = None
    raw_plan_output: str | None = None
    validation_warnings: list[str] = field(default_factory=list)
    coverage: dict[str, Any] | None = None


class TablePipelineService:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.storage = TableObjectStorage(settings)
        self.index = TableSearchIndex(settings)

    def ingest_table(
        self,
        filename: str,
        content: bytes,
        sheet_name: str | None = None,
        table_id: str | None = None,
        batch_id: str | None = None,
        source_object: str | None = None,
    ) -> dict[str, Any]:
        converted = convert_table_file_to_xlsx(filename, content)
        workbook = load_workbook(io.BytesIO(converted.xlsx_content), data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        parse_result = self._parse_sheet_for_pipeline(sheet)
        table_id = table_id or uuid4().hex
        batch_id = batch_id or table_id
        table_name = _table_display_name(parse_result.table_title, sheet.title, converted.original_filename)
        tree_with_name = _attach_table_name(parse_result.tree, table_name)
        tree_refs_with_name = _attach_table_name(parse_result.tree_with_cell_refs, table_name)
        sheet_normalized_filename = f"{_safe_sheet_filename(sheet.title)}.xlsx"
        sheet_xlsx_content = _workbook_sheet_to_xlsx(workbook, sheet.title)

        artifact = _json_safe(
            {
                "table_id": table_id,
                "batch_id": batch_id,
                "filename": converted.original_filename,
                "normalized_filename": sheet_normalized_filename,
                "source_extension": converted.source_extension,
                "sheet_name": sheet.title,
                "table_title": parse_result.table_title,
                "parse_mode": parse_result.parse_mode,
                "large_table_reason": parse_result.large_table_reason,
                "markdown_table": parse_result.markdown_table,
                "normalized_headers": parse_result.normalized_headers,
                "hierarchy_definition": parse_result.hierarchy_definition,
                "summary_text": parse_result.summary_text,
                "final_json_tree": parse_result.final_json_tree,
                "raw_plan_output": parse_result.raw_plan_output,
                "parse_plan": parse_result.parse_plan,
                "validation_warnings": parse_result.validation_warnings,
                "coverage": parse_result.coverage,
                "tree_with_cell_refs": tree_refs_with_name,
                "tree": tree_with_name,
            }
        )

        stored = self.storage.store_table_artifacts(
            table_id=table_id,
            batch_id=batch_id,
            source_filename=converted.original_filename,
            source_content=None if source_object else converted.source_content,
            source_object=source_object,
            normalized_filename=sheet_normalized_filename,
            xlsx_content=sheet_xlsx_content,
            artifact=artifact,
        )

        summary = build_table_summary(
            normalized_headers=parse_result.normalized_headers,
            hierarchy_definition=parse_result.hierarchy_definition,
            summary_text=parse_result.summary_text,
            tree=tree_with_name,
        )
        index_document = {
            "table_id": table_id,
            "batch_id": batch_id,
            "filename": converted.original_filename,
            "normalized_filename": sheet_normalized_filename,
            "source_extension": converted.source_extension,
            "sheet_name": sheet.title,
            "table_title": parse_result.table_title,
            "parse_mode": parse_result.parse_mode,
            "large_table_reason": parse_result.large_table_reason,
            "normalized_headers": parse_result.normalized_headers,
            "hierarchy_definition": parse_result.hierarchy_definition,
            "source_object": stored.source_object,
            "xlsx_object": stored.xlsx_object,
            "tree_object": stored.tree_object,
            **summary,
        }
        indexed_document = self.index.index_table(index_document)

        return _json_safe(
            {
                "table_id": table_id,
                "batch_id": batch_id,
                "filename": converted.original_filename,
                "normalized_filename": sheet_normalized_filename,
                "sheet_name": sheet.title,
                "table_title": parse_result.table_title,
                "parse_mode": parse_result.parse_mode,
                "large_table_reason": parse_result.large_table_reason,
                "coverage": parse_result.coverage,
                "summary_text": summary["summary_text"],
                "candidate_fields": summary["candidate_fields"],
                "minio_objects": {
                    "source_object": stored.source_object,
                    "xlsx_object": stored.xlsx_object,
                    "tree_object": stored.tree_object,
                },
                "indexed": True,
                "indexed_vector": bool(indexed_document.get("summary_vector")),
                "tree": tree_with_name,
            }
        )

    def _parse_sheet_for_pipeline(self, sheet: Any) -> PipelineParseResult:
        markdown_table = _sheet_to_markdown_for_size_check(sheet)
        large_table_reason = _large_table_reason(sheet, markdown_table, self.settings)
        table_title = _table_display_name(_extract_title_from_sheet(sheet), sheet.title)

        if large_table_reason:
            return self._parse_sheet_with_plan(
                sheet=sheet,
                markdown_table=markdown_table,
                table_title=table_title,
                large_table_reason=large_table_reason,
            )

        enhanced_parser = LangChainEnhancedTableParser(self.settings)
        enhanced_result = enhanced_parser.parse_sheet(sheet)
        quality_warnings = _validate_enhanced_tree_quality(enhanced_result.tree_with_cell_refs)
        if quality_warnings:
            fallback_reason = "enhanced_tree_quality_failed: " + "; ".join(quality_warnings[:5])
            print(
                "\n[ENHANCED_TREE_REJECTED]\n"
                f"reason={fallback_reason}\n"
                f"final_json_tree:\n{enhanced_result.final_json_tree}\n"
                "[/ENHANCED_TREE_REJECTED]\n",
                flush=True,
            )
            return self._parse_sheet_with_plan(
                sheet=sheet,
                markdown_table=markdown_table,
                table_title=table_title,
                large_table_reason=fallback_reason,
            )
        return PipelineParseResult(
            parse_mode="enhanced_llm",
            large_table_reason=None,
            table_title=enhanced_result.table_title,
            markdown_table=enhanced_result.markdown_table,
            normalized_headers=enhanced_result.normalized_headers,
            hierarchy_definition=enhanced_result.hierarchy_definition,
            summary_text=enhanced_result.summary_text,
            final_json_tree=enhanced_result.final_json_tree,
            tree_with_cell_refs=enhanced_result.tree_with_cell_refs,
            tree=enhanced_result.tree,
        )

    def _parse_sheet_with_plan(
        self,
        sheet: Any,
        markdown_table: str,
        table_title: str,
        large_table_reason: str,
    ) -> PipelineParseResult:
        plan_parser = PlanBasedTableParser(self.settings)
        plan_result = plan_parser.parse_sheet(sheet)
        parse_plan = plan_result.parse_plan.model_dump()
        coverage = plan_result.coverage.model_dump()
        coverage["is_complete"] = plan_result.coverage.is_complete
        normalized_headers = _plan_normalized_headers(parse_plan)
        hierarchy_definition = _plan_hierarchy_definition(
            parse_plan,
            plan_result.validation_warnings,
            coverage,
        )
        summary_text = _generate_summary_from_llm(
            settings=self.settings,
            markdown_table=markdown_table,
            normalized_headers=normalized_headers,
            hierarchy_definition=hierarchy_definition,
        )
        return PipelineParseResult(
            parse_mode="plan_based",
            large_table_reason=large_table_reason,
            table_title=table_title,
            markdown_table=markdown_table,
            normalized_headers=normalized_headers,
            hierarchy_definition=hierarchy_definition,
            summary_text=summary_text,
            final_json_tree=None,
            tree_with_cell_refs=plan_result.tree_with_cell_refs,
            tree=plan_result.tree,
            parse_plan=parse_plan,
            raw_plan_output=plan_result.raw_plan_output,
            validation_warnings=plan_result.validation_warnings,
            coverage=coverage,
        )

    def answer_question(
        self,
        question: str,
        top_k: int = 3,
        evidence_limit: int = 12,
        use_llm: bool = True,
    ) -> dict[str, Any]:
        qa_service = TableQAService(self.settings)
        llm = qa_service.llm if use_llm else None
        query_plan = _build_query_plan(question, llm)
        retrieval_questions = query_plan["sub_questions"]
        query_results = self._answer_retrieval_questions(
            retrieval_questions=retrieval_questions,
            top_k=top_k,
            evidence_limit=evidence_limit,
            use_llm=use_llm,
            qa_service=qa_service if len(retrieval_questions) == 1 else None,
        )
        table_candidates = _dedupe_query_candidates(query_results)
        table_answers = [
            item
            for result in query_results
            for item in result.get("table_answers", [])
        ]

        if not table_candidates:
            return _json_safe(
                {
                    "answer": "当前未检索到相关表格，无法回答问题。",
                    "mode": "pipeline_no_table_candidates",
                    "original_question": question,
                    "retrieval_question": (
                        retrieval_questions[0]
                        if len(retrieval_questions) == 1
                        else question
                    ),
                    "retrieval_questions": retrieval_questions,
                    "query_plan": query_plan,
                    "query_results": _trim_query_results(query_results),
                    "table_candidates": [],
                    "table_answers": [],
                }
            )

        answer = _synthesize_pipeline_answer(
            question=question,
            table_answers=table_answers,
            llm=llm,
        )
        return _json_safe(
            {
                "answer": answer,
                "mode": "es_minio_tree_qa",
                "original_question": question,
                "retrieval_question": (
                    retrieval_questions[0]
                    if len(retrieval_questions) == 1
                    else question
                ),
                "retrieval_questions": retrieval_questions,
                "query_plan": query_plan,
                "query_results": _trim_query_results(query_results),
                "table_candidates": _trim_candidate_payload(table_candidates),
                "table_answers": table_answers,
            }
        )

    def _answer_retrieval_questions(
        self,
        retrieval_questions: list[str],
        top_k: int,
        evidence_limit: int,
        use_llm: bool,
        qa_service: TableQAService | None = None,
    ) -> list[dict[str, Any]]:
        if len(retrieval_questions) <= 1:
            return [
                self._answer_single_retrieval_question(
                    retrieval_question=retrieval_questions[0],
                    top_k=top_k,
                    evidence_limit=evidence_limit,
                    use_llm=use_llm,
                    qa_service=qa_service,
                )
            ]

        results: list[dict[str, Any] | None] = [None] * len(retrieval_questions)
        worker_count = min(len(retrieval_questions), MAX_PARALLEL_SUB_QUERIES)
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            future_to_index = {
                executor.submit(
                    self._answer_single_retrieval_question,
                    retrieval_question=retrieval_question,
                    top_k=top_k,
                    evidence_limit=evidence_limit,
                    use_llm=use_llm,
                ): index
                for index, retrieval_question in enumerate(retrieval_questions)
            }
            for future in as_completed(future_to_index):
                results[future_to_index[future]] = future.result()
        return [result for result in results if result is not None]

    def _answer_single_retrieval_question(
        self,
        retrieval_question: str,
        top_k: int,
        evidence_limit: int,
        use_llm: bool,
        qa_service: TableQAService | None = None,
    ) -> dict[str, Any]:
        candidates = self.index.search(retrieval_question, top_k=top_k)
        if not candidates:
            return {
                "question": retrieval_question,
                "table_candidates": [],
                "table_answers": [
                    {
                        "sub_question": retrieval_question,
                        "retrieval_question": retrieval_question,
                        "table": None,
                        "error": "当前未检索到相关表格",
                    }
                ],
            }

        service = qa_service or TableQAService(self.settings)
        table_answers: list[dict[str, Any]] = []
        for candidate in candidates:
            table_payload = {
                "table_id": candidate.get("table_id"),
                "batch_id": candidate.get("batch_id"),
                "filename": candidate.get("filename"),
                "sheet_name": candidate.get("sheet_name"),
                "table_title": candidate.get("table_title"),
                "score": candidate.get("score"),
            }
            try:
                artifact = self.storage.get_json(candidate["tree_object"])
                metadata = {
                    "table_id": artifact.get("table_id"),
                    "filename": artifact.get("filename"),
                    "sheet_name": artifact.get("sheet_name"),
                    "table_title": artifact.get("table_title"),
                }
                qa_result = service.answer(
                    question=retrieval_question,
                    tree=artifact.get("tree") or {},
                    metadata=metadata,
                    limit=evidence_limit,
                    use_llm=use_llm,
                )
                table_answers.append(
                    {
                        "sub_question": retrieval_question,
                        "retrieval_question": retrieval_question,
                        "table": table_payload,
                        "qa": qa_result,
                    }
                )
            except Exception as exc:
                table_answers.append(
                    {
                        "sub_question": retrieval_question,
                        "retrieval_question": retrieval_question,
                        "table": table_payload,
                        "error": f"{exc.__class__.__name__}: {exc}",
                    }
                )

        return {
            "question": retrieval_question,
            "table_candidates": candidates,
            "table_answers": table_answers,
        }

    def get_table_artifact(self, table_id: str) -> dict[str, Any]:
        document = self.index.get_table_document(table_id)
        tree_object = document.get("tree_object") if document else None
        return self.storage.get_json(str(tree_object or self.storage.tree_object_name(table_id)))

    def get_table_summary(self, table_id: str) -> dict[str, Any]:
        artifact = self.get_table_artifact(table_id)
        document = self.index.get_table_document(table_id)
        return _json_safe(
            {
                "table_id": table_id,
                "batch_id": artifact.get("batch_id"),
                "filename": artifact.get("filename"),
                "normalized_filename": artifact.get("normalized_filename"),
                "source_extension": artifact.get("source_extension"),
                "sheet_name": artifact.get("sheet_name"),
                "table_title": artifact.get("table_title"),
                "parse_mode": artifact.get("parse_mode"),
                "large_table_reason": artifact.get("large_table_reason"),
                "coverage": artifact.get("coverage"),
                "minio_objects": artifact.get("minio_objects"),
                "summary_text": (document or {}).get("summary_text"),
                "candidate_fields": (document or {}).get("candidate_fields", []),
                "tree_metric_names": (document or {}).get("tree_metric_names", []),
                "embedding_error": (document or {}).get("embedding_error"),
                "indexed": document is not None,
            }
        )

    def list_table_summaries(self, limit: int = 50) -> list[dict[str, Any]]:
        tables: list[dict[str, Any]] = []
        for document in self.index.list_tables(limit=limit):
            tables.append(
                {
                    "table_id": document.get("table_id"),
                    "batch_id": document.get("batch_id"),
                    "filename": document.get("filename"),
                    "normalized_filename": document.get("normalized_filename"),
                    "source_extension": document.get("source_extension"),
                    "sheet_name": document.get("sheet_name"),
                    "table_title": document.get("table_title"),
                    "parse_mode": document.get("parse_mode"),
                    "large_table_reason": document.get("large_table_reason"),
                    "coverage": document.get("coverage"),
                    "summary_text": document.get("summary_text"),
                    "candidate_fields": document.get("candidate_fields", []),
                    "tree_metric_names": document.get("tree_metric_names", []),
                    "embedding_error": document.get("embedding_error"),
                    "created_at": document.get("created_at"),
                    "indexed": True,
                    "minio_objects": {
                        "source_object": document.get("source_object"),
                        "xlsx_object": document.get("xlsx_object"),
                        "tree_object": document.get("tree_object"),
                    },
                }
            )
        return _json_safe(tables)

    def get_table_tree(self, table_id: str) -> dict[str, Any]:
        artifact = self.get_table_artifact(table_id)
        table_name = _table_display_name(
            artifact.get("table_title"),
            artifact.get("sheet_name"),
            artifact.get("filename"),
        )
        return _json_safe(
            {
                "table_id": table_id,
                "batch_id": artifact.get("batch_id"),
                "filename": artifact.get("filename"),
                "sheet_name": artifact.get("sheet_name"),
                "table_title": artifact.get("table_title"),
                "tree": _attach_table_name(artifact.get("tree") or {}, table_name),
                "tree_with_cell_refs": _attach_table_name(
                    artifact.get("tree_with_cell_refs") or {},
                    table_name,
                ),
            }
        )

    def get_table_file(self, table_id: str, kind: str = "source") -> tuple[str, str, bytes]:
        artifact = self.get_table_artifact(table_id)
        objects = artifact.get("minio_objects") or {}
        if kind == "source":
            object_name = objects.get("source_object") or _lookup_object_from_index(
                self.index,
                table_id,
                "source_object",
            )
            filename = artifact.get("filename") or f"{table_id}-source"
            media_type = "application/octet-stream"
        elif kind == "normalized":
            object_name = objects.get("xlsx_object") or _lookup_object_from_index(
                self.index,
                table_id,
                "xlsx_object",
            )
            filename = artifact.get("normalized_filename") or f"{table_id}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ValueError(f"Unsupported table file kind: {kind}")

        if not object_name:
            raise ValueError(f"Cannot find MinIO object for table {table_id}: {kind}")
        return str(filename), media_type, self.storage.get_bytes(str(object_name))


def _build_query_plan(question: str, llm: Any | None) -> dict[str, Any]:
    normalized_question = question.strip()
    if not normalized_question:
        return {
            "original_question": question,
            "sub_questions": [question],
            "rewritten": False,
            "source": "original",
        }

    if llm is not None:
        try:
            sub_questions = _rewrite_queries_with_llm(normalized_question, llm)
            if sub_questions:
                return {
                    "original_question": normalized_question,
                    "sub_questions": sub_questions,
                    "rewritten": sub_questions != [normalized_question],
                    "source": "llm",
                }
        except Exception as exc:
            fallback = _split_question_heuristically(normalized_question)
            return {
                "original_question": normalized_question,
                "sub_questions": fallback,
                "rewritten": fallback != [normalized_question],
                "source": "heuristic_after_llm_error",
                "error": f"{exc.__class__.__name__}: {exc}",
            }

    sub_questions = _split_question_heuristically(normalized_question)
    return {
        "original_question": normalized_question,
        "sub_questions": sub_questions,
        "rewritten": sub_questions != [normalized_question],
        "source": "heuristic" if sub_questions != [normalized_question] else "original",
    }


def _rewrite_queries_with_llm(question: str, llm: Any) -> list[str]:
    prompt = f"""
你是一个表格问答 query 改写助手。请把用户 query 改写成适合独立检索的子问题。

规则：
1. 如果 query 同时询问多个并列对象、指标或统计项，请拆成多个完整子问题。
2. 每个子问题必须保留共同限定条件，例如学校、院系、年份、范围、单位、统计口径。
3. 每个子问题必须能单独作为检索 query 使用，不要输出只有名词的片段。
4. 如果 query 只有一个明确问题，请原样返回一个子问题。
5. 不要回答问题，只输出 JSON 对象。
6. sub_questions 最多 {MAX_REWRITTEN_QUERIES} 个。

示例：
用户 query：湖南师范大学的博士研究生，普通本科生，硕士研究生有多少
输出：{{"sub_questions":["湖南师范大学的博士研究生有多少","湖南师范大学的普通本科生有多少","湖南师范大学的硕士研究生有多少"]}}

输出格式：
{{"sub_questions":["完整子问题1","完整子问题2"]}}

用户 query：
{question}
"""
    response = llm.invoke([HumanMessage(content=prompt)])
    content = str(getattr(response, "content", response)).strip()
    payload = json.loads(_extract_json_object(content))
    values = payload.get("sub_questions", [])
    return _normalize_rewritten_queries(question, values)


def _split_question_heuristically(question: str) -> list[str]:
    stripped = question.strip()
    body = stripped.rstrip("？?")
    suffix = _detect_question_suffix(body)
    if not suffix:
        return [stripped]

    stem = body[: -len(suffix["matched"])].strip()
    parts = _split_parallel_parts(stem)
    if len(parts) < 2:
        return [stripped]

    prefix, first_item = _split_shared_prefix(parts[0])
    if not prefix or not first_item:
        return [stripped]

    rewritten: list[str] = []
    for index, part in enumerate(parts):
        item = first_item if index == 0 else part
        if not item:
            continue
        if "的" not in item and not item.startswith(prefix):
            item = f"{prefix}{item}"
        rewritten.append(f"{item}{suffix['normalized']}")
    return _normalize_rewritten_queries(stripped, rewritten) or [stripped]


def _detect_question_suffix(question: str) -> dict[str, str] | None:
    suffixes = [
        ("数量分别是多少", "数量是多少"),
        ("人数分别是多少", "人数是多少"),
        ("分别有多少", "有多少"),
        ("分别是多少", "是多少"),
        ("总数是多少", "总数是多少"),
        ("数量是多少", "数量是多少"),
        ("人数是多少", "人数是多少"),
        ("有多少", "有多少"),
        ("是多少", "是多少"),
        ("有哪些", "有哪些"),
        ("是什么", "是什么"),
    ]
    for matched, normalized in suffixes:
        if question.endswith(matched):
            return {"matched": matched, "normalized": normalized}
    return None


def _split_parallel_parts(text: str) -> list[str]:
    raw_parts = [
        part.strip()
        for part in re.split(r"[，,、；;]+", text)
        if part.strip()
    ]
    if len(raw_parts) < 2:
        return raw_parts

    parts: list[str] = []
    for part in raw_parts:
        sub_parts = [
            item.strip()
            for item in re.split(r"(?<!的)[和及与](?!的)", part)
            if item.strip()
        ]
        if len(sub_parts) > 1 and all("的" not in item for item in sub_parts):
            parts.extend(sub_parts)
        else:
            parts.append(part)
    return parts


def _split_shared_prefix(first_part: str) -> tuple[str, str]:
    separator_index = first_part.rfind("的")
    if separator_index < 0:
        return "", first_part
    prefix = first_part[: separator_index + 1].strip()
    item = first_part[separator_index + 1 :].strip()
    return prefix, item


def _normalize_rewritten_queries(question: str, values: Any) -> list[str]:
    if not isinstance(values, list):
        return []

    seen: set[str] = set()
    result: list[str] = []
    original_key = re.sub(r"\s+", "", question.lower())
    for value in values:
        text = re.sub(r"\s+", " ", str(value or "")).strip()
        text = re.sub(r"^[\d一二三四五六七八九十]+[.、)\）]\s*", "", text)
        if not text:
            continue
        key = re.sub(r"\s+", "", text.lower())
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
        if len(result) >= MAX_REWRITTEN_QUERIES:
            break
    if len(result) > 1:
        result = [
            text
            for text in result
            if re.sub(r"\s+", "", text.lower()) != original_key
        ] or result
    return result or [question]


def _dedupe_query_candidates(query_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    order: list[str] = []

    for query_result in query_results:
        question = query_result.get("question")
        for index, candidate in enumerate(query_result.get("table_candidates", [])):
            key = str(
                candidate.get("table_id")
                or candidate.get("tree_object")
                or f"{question}:{index}"
            )
            if key not in by_key:
                by_key[key] = {
                    **candidate,
                    "matched_queries": [question] if question else [],
                }
                order.append(key)
                continue

            existing = by_key[key]
            if _candidate_score(candidate) > _candidate_score(existing):
                existing.update(candidate)
            matched_queries = existing.setdefault("matched_queries", [])
            if question and question not in matched_queries:
                matched_queries.append(question)

    return [by_key[key] for key in order]


def _candidate_score(candidate: dict[str, Any]) -> float:
    score = candidate.get("score")
    return float(score) if isinstance(score, (int, float)) else float("-inf")


def _trim_query_results(query_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    trimmed: list[dict[str, Any]] = []
    for item in query_results:
        table_answers = item.get("table_answers", [])
        trimmed.append(
            {
                "question": item.get("question"),
                "table_candidates": _trim_candidate_payload(
                    item.get("table_candidates", [])
                ),
                "table_answer_count": len(table_answers),
                "answerable_count": len(_compact_answerable_table_answers(table_answers)),
            }
        )
    return trimmed


def _synthesize_pipeline_answer(
    question: str,
    table_answers: list[dict[str, Any]],
    llm: Any | None,
) -> str:
    compact_answers = _compact_answerable_table_answers(table_answers)
    compact_answers.extend(_compact_unanswered_sub_questions(table_answers, compact_answers))
    if not compact_answers:
        compact_answers = [
            {
                "table": item.get("table"),
                "sub_question": item.get("sub_question"),
                "answer": item.get("qa", {}).get("answer"),
                "evidence_paths": item.get("qa", {}).get("evidence_paths", []),
                "error": item.get("error"),
            }
            for item in table_answers
        ]
    if llm is None:
        lines = ["检索到的表格问答结果："]
        for item in compact_answers:
            table = item.get("table") or {}
            title = table.get("table_title") or table.get("filename") or table.get("table_id")
            sub_question = item.get("sub_question")
            prefix = f"{sub_question} - " if sub_question else ""
            lines.append(f"- {prefix}{title}: {item.get('answer') or item.get('error')}")
        return "\n".join(lines)

    prompt = f"""
你是一个跨表格问答助手。请只根据 table_answers 中的答案和 evidence_paths 综合回答用户问题。

要求：
1. 不要使用外部知识。
2. 如果证据足够，直接回答，并说明来自哪些表格。
3. 如果多个表格给出互补证据，请合并回答。
4. 如果候选表格都没有足够证据，请说明“当前证据不足”。
5. 如果 table_answers 中包含 sub_question，请逐个覆盖这些子问题，再汇总回答原始问题。

用户问题：
{question}

table_answers:
{json.dumps(compact_answers, ensure_ascii=False, indent=2, default=str)}
"""
    response = llm.invoke([HumanMessage(content=prompt)])
    return str(getattr(response, "content", response)).strip()


def _compact_answerable_table_answers(table_answers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    compact: list[dict[str, Any]] = []
    for item in table_answers:
        qa = item.get("qa") or {}
        evidence_paths = qa.get("evidence_paths", [])
        answer = str(qa.get("answer") or "")
        if not evidence_paths:
            continue
        if "当前证据不足" in answer:
            continue
        compact.append(
            {
                "table": item.get("table"),
                "sub_question": item.get("sub_question"),
                "answer": answer,
                "evidence_paths": evidence_paths,
                "error": item.get("error"),
            }
        )
    return compact


def _compact_unanswered_sub_questions(
    table_answers: list[dict[str, Any]],
    compact_answers: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    all_sub_questions = _dedupe_strings(
        [
            str(item.get("sub_question") or "")
            for item in table_answers
            if item.get("sub_question")
        ]
    )
    if len(all_sub_questions) <= 1:
        return []

    answered = {
        str(item.get("sub_question") or "")
        for item in compact_answers
        if item.get("sub_question")
    }
    unanswered: list[dict[str, Any]] = []
    for sub_question in all_sub_questions:
        if sub_question in answered:
            continue
        related = [
            item
            for item in table_answers
            if item.get("sub_question") == sub_question
        ]
        first = related[0] if related else {}
        qa = first.get("qa") or {}
        unanswered.append(
            {
                "table": first.get("table"),
                "sub_question": sub_question,
                "answer": qa.get("answer") or first.get("error") or "当前证据不足",
                "evidence_paths": qa.get("evidence_paths", []),
                "error": first.get("error"),
            }
        )
    return unanswered


def _trim_candidate_payload(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for candidate in candidates:
        result.append(
            {
                "score": candidate.get("score"),
                "table_id": candidate.get("table_id"),
                "batch_id": candidate.get("batch_id"),
                "filename": candidate.get("filename"),
                "sheet_name": candidate.get("sheet_name"),
                "table_title": candidate.get("table_title"),
                "tree_object": candidate.get("tree_object"),
                "tree_metric_names": candidate.get("tree_metric_names", [])[:80],
                "matched_queries": candidate.get("matched_queries", []),
            }
        )
    return result


def _extract_json_object(text: str) -> str:
    if "```json" in text:
        return text.split("```json", 1)[1].split("```", 1)[0].strip()
    if "```" in text:
        return text.split("```", 1)[1].split("```", 1)[0].strip()
    start = text.find("{")
    end = text.rfind("}") + 1
    if start < 0 or end <= start:
        raise ValueError("No JSON object found")
    return text[start:end]


def _dedupe_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = re.sub(r"\s+", "", value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(value)
    return result


def _lookup_object_from_index(index: TableSearchIndex, table_id: str, field_name: str) -> Any:
    document = index.get_table_document(table_id)
    if not document:
        return None
    return document.get(field_name)


def _validate_enhanced_tree_quality(tree_with_cell_refs: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    leaf_count = 0
    bad_leaf_count = 0
    bad_examples: list[str] = []
    suspicious_examples: list[str] = []
    suspicious_patterns = [
        "极简输出",
        "仅展示",
        "完整输出",
        "如需",
        "结构示例",
        "模式已启用",
        "省略",
        "未完",
        "truncated",
    ]

    def walk(value: Any, path: list[str]) -> None:
        nonlocal leaf_count, bad_leaf_count
        if isinstance(value, dict):
            for key, item in value.items():
                walk(item, [*path, str(key)])
            return
        if isinstance(value, list):
            for index, item in enumerate(value):
                walk(item, [*path, str(index)])
            return

        leaf_count += 1
        text = str(value)
        if any(pattern in text for pattern in suspicious_patterns):
            if len(suspicious_examples) < 3:
                suspicious_examples.append(f"{' | '.join(path)}={text[:120]}")
        if not _is_cell_or_range_ref(text):
            bad_leaf_count += 1
            if len(bad_examples) < 3:
                bad_examples.append(f"{' | '.join(path)}={text[:120]}")

    walk(tree_with_cell_refs, [])
    if leaf_count == 0:
        warnings.append("no_leaf_values")
    if suspicious_examples:
        warnings.append("model_comment_in_tree: " + " || ".join(suspicious_examples))
    if bad_leaf_count:
        warnings.append(
            f"non_cell_ref_leaf_values={bad_leaf_count}/{leaf_count}: "
            + " || ".join(bad_examples)
        )
    return warnings


def _is_cell_or_range_ref(value: str) -> bool:
    text = value.strip().upper()
    return bool(
        re.fullmatch(r"\$?[A-Z]+\$?[0-9]+(:\$?[A-Z]+\$?[0-9]+)?", text)
    )


def _large_table_reason(sheet: Any, markdown_table: str, settings: Settings) -> str | None:
    reasons: list[str] = []
    cell_count = int(sheet.max_row or 0) * int(sheet.max_column or 0)
    merged_cells = getattr(sheet, "merged_cells", None)
    merged_ranges = getattr(merged_cells, "ranges", []) if merged_cells is not None else []
    merged_count = len(merged_ranges)
    if cell_count > settings.large_table_cell_threshold:
        reasons.append(f"cells={cell_count}>{settings.large_table_cell_threshold}")
    if int(sheet.max_row or 0) > settings.large_table_row_threshold:
        reasons.append(f"rows={sheet.max_row}>{settings.large_table_row_threshold}")
    if int(sheet.max_column or 0) > settings.large_table_column_threshold:
        reasons.append(f"columns={sheet.max_column}>{settings.large_table_column_threshold}")
    if merged_count > settings.large_table_merged_cell_threshold:
        reasons.append(f"merged={merged_count}>{settings.large_table_merged_cell_threshold}")
    if len(markdown_table) > settings.large_table_markdown_threshold:
        reasons.append(f"markdown={len(markdown_table)}>{settings.large_table_markdown_threshold}")
    return "; ".join(reasons) if reasons else None


def _sheet_to_markdown_for_size_check(sheet: Any) -> str:
    from app.services.table2tree_enhanced import excel_to_markdown_with_cell_ref

    return excel_to_markdown_with_cell_ref(sheet)


def _extract_title_from_sheet(sheet: Any) -> str:
    from app.services.table2tree_enhanced import extract_table_title

    return extract_table_title(sheet)


def _plan_normalized_headers(parse_plan: dict[str, Any]) -> str:
    headers = {
        "hierarchy_columns": parse_plan.get("hierarchy_columns", []),
        "value_columns": parse_plan.get("value_columns", []),
    }
    return json.dumps(headers, ensure_ascii=False)


def _plan_hierarchy_definition(
    parse_plan: dict[str, Any],
    validation_warnings: list[str],
    coverage: dict[str, Any],
) -> str:
    payload = {
        "table_range": parse_plan.get("table_range"),
        "title_ranges": parse_plan.get("title_ranges", []),
        "header_ranges": parse_plan.get("header_ranges", []),
        "data_row_range": parse_plan.get("data_row_range"),
        "hierarchy_columns": parse_plan.get("hierarchy_columns", []),
        "value_columns": parse_plan.get("value_columns", []),
        "hierarchy_fill_down": parse_plan.get("hierarchy_fill_down"),
        "validation_warnings": validation_warnings,
        "coverage": coverage,
    }
    return json.dumps(payload, ensure_ascii=False)


def _generate_summary_from_llm(
    settings: Settings,
    markdown_table: str,
    normalized_headers: str,
    hierarchy_definition: str,
) -> str:
    if not settings.llm_model or not settings.llm_api_key:
        return ""
    kwargs: dict[str, Any] = {
        "model": settings.llm_model,
        "api_key": settings.llm_api_key,
        "temperature": settings.llm_temperature,
        "timeout": settings.llm_timeout_seconds,
    }
    if settings.llm_base_url:
        kwargs["base_url"] = settings.llm_base_url
    llm = ChatOpenAI(**kwargs)
    prompt = ChatPromptTemplate.from_template(TABLE_SUMMARY_PROMPT)
    response = (prompt | llm).invoke(
        {
            "TABLE_AS_JSON_STRING": markdown_table,
            "NORMALIZED_HEADERS_FROM_STEP_1": normalized_headers,
            "HIERARCHY_DEFINITION_FROM_STEP_2": hierarchy_definition,
        }
    )
    return str(getattr(response, "content", response)).strip()


def _table_display_name(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return "未命名表格"


def _attach_table_name(tree: Any, table_name: str) -> dict[str, Any]:
    if isinstance(tree, dict):
        if tree.get("表格名") == table_name:
            return tree
        return {
            "表格名": table_name,
            **{key: value for key, value in tree.items() if key != "表格名"},
        }
    return {
        "表格名": table_name,
        "数据": tree,
    }


def _workbook_sheet_to_xlsx(workbook: Workbook, sheet_name: str) -> bytes:
    source_sheet = workbook[sheet_name]
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = _safe_excel_sheet_title(sheet_name)

    for row in source_sheet.iter_rows():
        for cell in row:
            output_sheet[cell.coordinate].value = cell.value

    output = io.BytesIO()
    output_workbook.save(output)
    return output.getvalue()


def _safe_excel_sheet_title(sheet_name: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", str(sheet_name).strip()) or "Sheet1"
    return title[:31]


def _safe_sheet_filename(sheet_name: str) -> str:
    cleaned = re.sub(r"[^\w.\-()\u4e00-\u9fff]+", "_", str(sheet_name).strip())
    return cleaned or "Sheet1"


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
